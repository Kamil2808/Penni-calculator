import tkinter.messagebox
from tkinter import *
from tkinter import ttk
import openpyxl
from openpyxl import Workbook, load_workbook
import math

wb = load_workbook('data.xlsx')
ws = wb.active

window = Tk()
window.title("Калькулятор пеней")
window.geometry('1280x720')
window.iconbitmap(r'icon.ico')

def clicked():
    delen = 0
    otvet = 0
    day_after = txt.get()
    day_after1 = int(day_after) - 0
    month_after = txt1.get()
    month_after1 = int(month_after) - 0
    year_after = txt2.get()
    year_after1 = int(year_after) - 2000
    year_after2 = int(year_after) / 4
    day_before = txt4.get()
    day_before1 = int(day_before) - 0
    month_before = txt5.get()
    month_before1 = int(month_before) - 0
    year_before = txt6.get()
    year_before1 = int(year_before) - 2000
    year_before2 = int(year_before) / 4
    summa_prosroch = txt3.get()
    stavka = txt7.get()
    lbl13.configure(text=summa_prosroch)
    lbl14.configure(text=day_after)
    lbl15.configure(text=month_after)
    lbl16.configure(text=year_after)
    if math.floor(int(year_after2)) == year_after2 and int(month_after) > 2:
        ws['A3'].value = 1
    else:
        ws['A3'].value = 0
    if math.floor(int(year_before2)) == year_before2 and int(month_before) > 2:
        ws['B3'].value = 1
    else:
        ws['B3'].value = 0
    if month_after1 == 1:
        ws['A1'].value = (day_after1)
    if month_after1 == 2:
        ws['A1'].value = (day_after1 + 28)
    if month_after1 == 3:
        ws['A1'].value = (day_after1 + 59)
    if month_after1 == 4:
        ws['A1'].value = (day_after1 + 89)
    if month_after1 == 5:
        ws['A1'].value = (day_after1 + 120)
    if month_after1 == 6:
        ws['A1'].value = (day_after1 + 150)
    if month_after1 == 7:
        ws['A1'].value = (day_after1 + 181)
    if month_after1 == 8:
        ws['A1'].value = (day_after1 + 212)
    if month_after1 == 9:
        ws['A1'].value = (day_after1 + 242)
    if month_after1 == 10:
        ws['A1'].value = (day_after1 + 273)
    if month_after1 == 11:
        ws['A1'].value = (day_after1 + 303)
    if month_after1 == 12:
        ws['A1'].value = (day_after1 + 334)
    if month_before1 == 1:
        ws['B1'].value = (day_before1)
    if month_before1 == 2:
        ws['B1'].value = (day_before1 + 28)
    if month_before1 == 3:
        ws['B1'].value = (day_before1 + 59)
    if month_before1 == 4:
        ws['B1'].value = (day_before1 + 89)
    if month_before1 == 5:
        ws['B1'].value = (day_before1 + 120)
    if month_before1 == 6:
        ws['B1'].value = (day_before1 + 150)
    if month_before1 == 7:
        ws['B1'].value = (day_before1 + 181)
    if month_before1 == 8:
        ws['B1'].value = (day_before1 + 212)
    if month_before1 == 9:
        ws['B1'].value = (day_before1 + 242)
    if month_before1 == 10:
        ws['B1'].value = (day_before1 + 273)
    if month_before1 == 11:
        ws['B1'].value = (day_before1 + 303)
    if month_before1 == 12:
        ws['B1'].value = (day_before1 + 334)
    ws['A2'].value = (year_after1 * 365)
    ws['B2'].value = (year_before1 * 365)
    ws['A1'].value = ws['A1'].value + ws['A2'].value + ws['A3'].value + 1
    ws['B1'].value = ws['B1'].value + ws['B2'].value + ws['B3'].value
    ws['C1'].value = ws['A1'].value - ws['B1'].value
    lbl18.configure(text=ws['C1'].value)
    if ws['C1'].value <= 60:
        delen = 300
    if ws['C1'].value > 60 <= 90:
        delen = 170
    if ws['C1'].value > 90:
        delen = 130
    otvet = int(summa_prosroch) * ws['C1'].value * float(stavka) / int(delen)
    otvet_okr = round(otvet, 2)
    lbl21.configure(text=otvet_okr)
    wb.save("data.xlsx")

lbl = Label(window, text="Дата оконча")
lbl.grid(column=0, row=0)
lbl1 = Label(window, text="ния расчетн")
lbl1.grid(column=1, row=0)
lbl2 = Label(window, text="ого периуда")
lbl2.grid(column=2, row=0)
lbl3 = Label(window, text="День")
lbl3.grid(column=0, row=2)
lbl4 = Label(window, text="Месяц")
lbl4.grid(column=1, row=2)
lbl5 = Label(window, text="Год")
lbl5.grid(column=2, row=2)
txt = Spinbox(window, from_=1, to=31, width=11)
txt.grid(column=0, row=1)
txt1 = Spinbox(window, from_=1, to=12, width=11)
txt1.grid(column=1, row=1)
txt2 = Spinbox(window, from_=2000, to=99999, width=11)
txt2.grid(column=2, row=1)
lbl6 = Label(window, text="Сумма просрочки")
lbl6.grid(column=3, row=0)
txt3 = Spinbox(window, from_=0, to=9999999999999999999999999999999999, width=17)
txt3.grid(column=3, row=1)
lbl6 = Label(window, text="Сумма просрочки")
lbl6.grid(column=4, row=0)
lbl13 = Label(window, text="")
lbl13.grid(column=4, row=1)
lbl7 = Label(window, text="Дата начал")
lbl7.grid(column=5, row=0)
lbl8 = Label(window, text="а расчетно")
lbl8.grid(column=6, row=0)
lbl9 = Label(window, text="го периуда")
lbl9.grid(column=7, row=0)
lbl10 = Label(window, text="День")
lbl10.grid(column=5, row=2)
lbl11 = Label(window, text="Месяц")
lbl11.grid(column=6, row=2)
lbl12 = Label(window, text="Год")
lbl12.grid(column=7, row=2)
txt4 = Spinbox(window, from_=1, to=31, width=8)
txt4.grid(column=5, row=1)
txt5 = Spinbox(window, from_=1, to=12, width=8)
txt5.grid(column=6, row=1)
txt6 = Spinbox(window, from_=2000, to=99999, width=8)
txt6.grid(column=7, row=1)
lbl = Label(window, text="Дата оконча")
lbl.grid(column=8, row=0)
lbl1 = Label(window, text="ния расчетн")
lbl1.grid(column=9, row=0)
lbl2 = Label(window, text="ого периуда")
lbl2.grid(column=10, row=0)
lbl14 = Label(window, text="")
lbl14.grid(column=8, row=1)
lbl15 = Label(window, text="")
lbl15.grid(column=9, row=1)
lbl16 = Label(window, text="")
lbl16.grid(column=10, row=1)
lbl17 = Label(window, text="Дней просрочки")
lbl17.grid(column=14, row=0)
lbl18 = Label(window, text="")
lbl18.grid(column=14, row=1)
lbl19 = Label(window, text="Ставка")
lbl19.grid(column=15, row=0)
txt7 = Spinbox(window, from_=0, to=100, width=8)
txt7.grid(column=15, row=1)
lbl20 = Label(window, text="Сумма процентов")
lbl20.grid(column=16, row=0)
lbl21 = Label(window, text="")
lbl21.grid(column=16, row=1)
btn = Button(window, text="Посчитать", command=clicked)
btn.grid(column=0, row=6)
inv1 = Label(window, text="")
inv1.grid(column=0, row=5)
window.mainloop()

ws.delete_cols(0, 100)
wb.save("data.xlsx")